import math
import json
import os
import streamlit as st
import pandas as pd
import requests
import hashlib
import concurrent.futures
import time
from io import BytesIO
from datetime import datetime, timedelta

# ----------------------------
# Constants
# ----------------------------
CC_CLIENT_ID     = "ae507531-707f-4bd1-9eb0-ae6685b01e6a"
CC_CLIENT_SECRET = "aMVpumtSXDF7LXhyo1UAFg"
CC_TOKEN_URL     = "https://identity.constantcontact.com/oauth2/aus1lm3ry9mF7x2Ja0h8/v1/token"

MC_CACHE_PATH        = os.path.expanduser("~/.vip_report_mc_cache.json")
MC_CACHE_REPO_PATH   = os.path.join(os.path.dirname(__file__), "mc_cache.json")

# MC: bulk/engagement segment IDs to EXCLUDE
MC_BULK_SEGMENT_IDS = {7299631}

# CC: all known tag IDs → display names (184 tags)
CC_TAG_NAME_MAP = {
    "01f1836c-2030-11ee-82a6-fa163e6b8330": "Senate 29",
    "0321743c-8c69-11ed-b794-fa163ed70180": "District 96",
    "03afcd96-8c63-11ed-a5e6-fa163e761ca9": "District 31",
    "06930d74-9759-11ed-a088-fa163e5fbd10": "District 45",
    "07db5b06-8c68-11ed-9e0e-fa163e5fbd10": "District 133",
    "084d7f7a-8c67-11ed-93f4-fa163e33fba4": "District 137",
    "0cf2019c-8c60-11ed-be1e-fa163e7e5464": "Senate 15",
    "0e1a747e-669d-11ee-9e16-fa163e6b8330": "PSC",
    "0e6289d8-1ff0-11ee-95cf-fa163ec0786c": "Senate 60",
    "0f3d447a-8c65-11ed-b53a-fa163e5fbd10": "District 63",
    "0f9d51f0-9758-11ed-ae2d-fa163e6b8330": "District 20",
    "104bf9ce-975a-11ed-8339-fa163e6b8330": "Assembly 64",
    "1098952a-8c62-11ed-bf30-fa163ef7d836": "Senate 21",
    "1103ec16-2033-11ee-84fd-fa163e5fbd10": "Assembly 23",
    "1105804a-98fb-11ed-9299-fa163ef7d836": "Senate 9",
    "11120276-2307-11ed-a126-fa163e5fbd10": "Climate Action Council",
    "11f3aa0c-8c64-11ed-ae2d-fa163e6b8330": "District 83",
    "125b337c-8c67-11ed-ab1d-fa163e761ca9": "District 122",
    "125c2ae2-8c68-11ed-a8dd-fa163e761ca9": "District 84",
    "129e3b70-8c69-11ed-b543-fa163ed70180": "District 119",
    "167611ee-8c61-11ed-84c5-fa163e5fbd10": "Senate 28",
    "19af6672-8c65-11ed-85e9-fa163ec0786c": "District 72",
    "1b983a50-bba6-11ef-9a05-fa163e02ac66": "Gillibrand",
    "1cc604b0-8c60-11ed-ad98-fa163e5fbd10": "Senate 36",
    "215e6ade-8c66-11ed-93c9-fa163ed70180": "District 139",
    "21bc3c2e-43d5-11ed-be87-fa163e7e5464": "NYC Casino Siting Board",
    "231a4e16-8d5f-11ed-93c9-fa163ed70180": "Senate 47",
    "2480a6f2-3c3b-11ee-853e-fa163e6b8330": "NY State Gaming Commissioner",
    "29bda934-8c65-11ed-9e0e-fa163e5fbd10": "District 70",
    "30ace91c-8c60-11ed-be2e-fa163ec0786c": "Senate 5",
    "31798132-9757-11ed-b7db-fa163e33fba4": "District 2",
    "33d8699a-8c65-11ed-a5e6-fa163e761ca9": "District 54",
    "341dc354-98fa-11ed-9a19-fa163e7e5464": "District 111",
    "355d0868-8c63-11ed-aa8b-fa163ef7d836": "District 107",
    "35caf262-8c66-11ed-ab1d-fa163e761ca9": "District 128",
    "36d73eb8-8c61-11ed-9299-fa163ef7d836": "Senate 7",
    "3c8c8134-017c-11ee-b38a-fa163ed70180": "Media - LCA",
    "3f53c9fa-9759-11ed-93c9-fa163ed70180": "District 50",
    "408961e2-2032-11ee-9c34-fa163e33fba4": "District 123",
    "41dc2ab4-8c5f-11ed-8128-fa163e6b8330": "Senate 6",
    "41ebb8f6-8c66-11ed-b543-fa163ed70180": "District 29",
    "420b5cae-a380-11f0-9329-fa163eddb7c0": "Mamdani",
    "42b4e7b4-8c64-11ed-89af-fa163e7e5464": "District 76",
    "432e71e8-8c68-11ed-be2e-fa163ec0786c": "District 8",
    "444b9738-8c61-11ed-89af-fa163e7e5464": "Senate 50",
    "44935e68-8c63-11ed-8f23-fa163e33fba4": "District 35",
    "44cb1d74-975a-11ed-ab1d-fa163e761ca9": "District 73",
    "44d67528-8c67-11ed-a8ac-fa163e7e5464": "District 51",
    "45b01ad2-8c65-11ed-b543-fa163ed70180": "District 78",
    "4ea46fea-8c64-11ed-89c1-fa163e33fba4": "District 12",
    "4ed76e0a-8c68-11ed-ae2d-fa163e6b8330": "District 127",
    "4f551c76-8c66-11ed-9524-fa163ec0786c": "District 79",
    "5215e526-975b-11ed-a7e2-fa163e7e5464": "District 101",
    "5321194a-98fa-11ed-a8dd-fa163e761ca9": "District 142",
    "53f24fe8-8c65-11ed-8339-fa163e6b8330": "District 48",
    "5631e3a0-8c5f-11ed-84c5-fa163e5fbd10": "Senate 30",
    "575fd0f0-8c61-11ed-86e7-fa163e7e5464": "Senate 37",
    "58aadbb6-9757-11ed-bece-fa163ec0786c": "District 4",
    "58d41642-8c68-11ed-b3f5-fa163e5fbd10": "District 71",
    "58ee3226-5305-11ed-a8ee-fa163ec0786c": "media",
    "58f6d45e-8c66-11ed-adb0-fa163e761ca9": "District 11",
    "59aeef54-8c60-11ed-86ea-fa163e33fba4": "Senate 32",
    "59f5fb34-9759-11ed-a088-fa163e5fbd10": "District 52",
    "5ae8af70-8c67-11ed-bb1e-fa163e6b8330": "District 65",
    "5e24b790-2034-11ee-af81-fa163e33fba4": "Assembly 76",
    "61372638-8c65-11ed-93f4-fa163e33fba4": "District 74",
    "628ac430-8c66-11ed-9524-fa163ec0786c": "District 134",
    "6413adc6-8c61-11ed-ae2d-fa163e6b8330": "Senate 12",
    "642e0d80-8c5f-11ed-8f23-fa163e33fba4": "Senate 23",
    "648aca4c-edaf-11ec-82ae-fa163e5fbd10": "Mayor Adams",
    "6492855a-8c62-11ed-89af-fa163e7e5464": "Senate 2",
    "651e8b2a-202f-11ee-b4df-fa163e6b8330": "Senate 51",
    "658c759a-8c64-11ed-9299-fa163ef7d836": "District 85",
    "66224266-8c67-11ed-be2e-fa163ec0786c": "District 69",
    "6982051e-98fa-11ed-ab1d-fa163e761ca9": "District 144",
    "6bb0bed8-2d53-11ed-af20-fa163ec0786c": "Traffic Mobility Review board",
    "6d4b1738-9758-11ed-bece-fa163ec0786c": "District 28",
    "6d8a406e-8c65-11ed-ab1d-fa163e761ca9": "District 109",
    "6eb80be4-8c63-11ed-9524-fa163ec0786c": "District 140",
    "6edad42e-3819-11ed-abc4-fa163e33fba4": "Hochul",
    "6f7325ca-8c61-11ed-b543-fa163ed70180": "Senate 41",
    "6fab9fbe-8c60-11ed-bb1e-fa163e6b8330": "Senate 42",
    "70672f52-8c62-11ed-aa8b-fa163ef7d836": "Senate 11",
    "70a57410-8c67-11ed-ab1d-fa163e761ca9": "District 91",
    "722737e8-8c66-11ed-a5e6-fa163e7e5464": "District 115",
    "739f3e20-8c5e-11ed-8128-fa163e6b8330": "Senate 22",
    "74f1afd2-8c5f-11ed-8e03-fa163e6b8330": "Senate 61",
    "76136ecc-8c65-11ed-882a-fa163e6b8330": "District 61",
    "78ff3130-9759-11ed-a8dd-fa163e761ca9": "District 57",
    "794b8444-b9de-11ec-bd8a-fa163ec0786c": "VIP",
    "79551408-8c67-11ed-b7db-fa163e33fba4": "District 132",
    "7b255484-975a-11ed-9a19-fa163e7e5464": "District 82",
    "7cdf60f0-8c63-11ed-8571-fa163ef7d836": "District 106",
    "7d089652-8c61-11ed-a7e2-fa163e7e5464": "Senate 4",
    "7d3a0590-8c64-11ed-ae2d-fa163e6b8330": "District 103",
    "7d973b3e-202f-11ee-87cb-fa163ed70180": "Senate 1",
    "7e41d4be-f48a-11ef-96c4-fa163e5a0a14": "Cuomo",
    "81ac9298-8c67-11ed-9e0e-fa163e5fbd10": "District 88",
    "81f21fd4-8c62-11ed-ab1d-fa163e761ca9": "Senate 20",
    "8217d3b0-a901-11ef-909d-fa163e213b1a": "Ritchie Torres",
    "828157a6-8c5f-11ed-a8dd-fa163e761ca9": "Senate 38",
    "89aff074-8c63-11ed-adb0-fa163e761ca9": "District 42",
    "8ab862d8-8c65-11ed-93f4-fa163e33fba4": "District 80",
    "8d1b6348-9757-11ed-adb0-fa163e761ca9": "District 15",
    "8e3f9ec0-8c66-11ed-93c9-fa163ed70180": "District 22",
    "8ed67906-8c5e-11ed-84c5-fa163e5fbd10": "Senate 34",
    "932afda8-8c61-11ed-86ea-fa163e33fba4": "Senate 48",
    "94a1aa68-9758-11ed-b7db-fa163e33fba4": "District 30",
    "95896e12-8c67-11ed-a7e2-fa163e7e5464": "District 38",
    "971757bc-98fb-11ed-b796-fa163e33fba4": "Senate 43",
    "9985d2e2-8c5f-11ed-9524-fa163ec0786c": "Senate 8",
    "9a1f86ac-8c65-11ed-b53a-fa163e5fbd10": "District 124",
    "9b7e8fb0-8c66-11ed-85e9-fa163ec0786c": "District 125",
    "9bb4dd8e-8c68-11ed-8e03-fa163e6b8330": "District 18",
    "9c1571a2-9759-11ed-9ab4-fa163e6b8330": "District 58",
    "9d0d832e-8c64-11ed-8d1e-fa163ed70180": "District 77",
    "9fc0323a-8c67-11ed-9adb-fa163ec0786c": "District 19",
    "a03a33a6-975b-11ed-a7e2-fa163e7e5464": "District 105",
    "a077b406-363e-11ee-9536-fa163e6b8330": "Gaming",
    "a095e200-8c61-11ed-bb1e-fa163e6b8330": "Senate 54",
    "a389df3c-8c5e-11ed-86ea-fa163e33fba4": "Senate 46",
    "a49d1376-8c68-11ed-b7db-fa163e33fba4": "District 33",
    "a5d2b436-8c62-11ed-9524-fa163ec0786c": "District 16",
    "a6680208-8c66-11ed-bb5c-fa163e7e5464": "District 40",
    "a72319e4-8b0b-11ef-8c9a-fa163e45f8ff": "AG James",
    "a7c53cc6-8c5f-11ed-b794-fa163ed70180": "Senate 44",
    "a982393e-8c63-11ed-b53a-fa163e5fbd10": "District 136",
    "abed4072-8c65-11ed-b543-fa163ed70180": "District 46",
    "ac86a232-8c68-11ed-b794-fa163ed70180": "District 55",
    "aff9ab02-8c64-11ed-9299-fa163ef7d836": "District 44",
    "b1160128-8c61-11ed-ae2d-fa163e6b8330": "Senate 40",
    "b1d5e58c-8c67-11ed-8128-fa163e6b8330": "District 9",
    "b262efc4-8c60-11ed-a8dd-fa163e761ca9": "Senate 13",
    "b333f912-8c5f-11ed-9959-fa163e33fba4": "Senate 33",
    "b5044f36-8c68-11ed-ad98-fa163e5fbd10": "District 143",
    "b5520d4c-9759-11ed-9a19-fa163e7e5464": "District 59",
    "b559a302-8c66-11ed-93f4-fa163e33fba4": "District 131",
    "b60439fc-98fa-11ed-b7db-fa163e33fba4": "District 147",
    "b7b603da-2119-11f0-bdbb-fa163e02ac66": "Grace Linnea Project",
    "b7d77f36-2032-11ee-8d9b-fa163e7e5464": "District 138",
    "b923d032-8c5e-11ed-9959-fa163e33fba4": "Senate 27",
    "bac21dfc-8c65-11ed-9ab4-fa163e6b8330": "District 7",
    "bb475e76-975b-11ed-b522-fa163ef7d836": "District 110",
    "bbf7b4aa-8c67-11ed-89af-fa163e7e5464": "District 6",
    "bc1fbc9a-1fef-11ee-9368-fa163e7e5464": "Senate 57",
    "bcce9564-8c68-11ed-a8dd-fa163e761ca9": "District 112",
    "bd818816-975a-11ed-a03e-fa163ec0786c": "District 93",
    "bf403902-2032-11ee-8ee5-fa163ef7d836": "District 81",
    "bf44f89e-b5cb-11f0-91fd-0242f225fc48": "Elise",
    "c252a0aa-8c60-11ed-8ce1-fa163ed70180": "Senate 10",
    "c3a982f6-8c66-11ed-be3f-fa163e5fbd10": "District 135",
    "c49ec3a4-8c68-11ed-a5e6-fa163e7e5464": "District 41",
    "c680cd08-8c67-11ed-8e03-fa163e6b8330": "District 27",
    "c69e93c6-8c65-11ed-9ab4-fa163e6b8330": "District 66",
    "c9ae7762-8c5f-11ed-a8dd-fa163e761ca9": "Senate 16",
    "ccbe5906-8c62-11ed-ae2d-fa163e6b8330": "District 21",
    "ce07c34e-8c60-11ed-aa8b-fa163ef7d836": "Senate 18",
    "ce90d22c-8c67-11ed-8e03-fa163e6b8330": "District 67",
    "cf8cfb70-8c61-11ed-8f23-fa163e33fba4": "Senate 62",
    "d0538e8a-8c65-11ed-ae2d-fa163e6b8330": "District 34",
    "d5422c0a-1c0c-11f1-9405-02427164ff99": "Heastie",
    "d5e21fda-8c5e-11ed-995d-fa163e6b8330": "Senate 26",
    "d8c58e9e-9758-11ed-adb0-fa163e761ca9": "District 37",
    "d8c86d42-8212-11f0-a2e5-fa163e5a0a14": "Rita",
    "d99d3c7e-eab2-11ed-bf29-fa163e5fbd10": "Borough President",
    "dadb52d0-8c5f-11ed-8f23-fa163e33fba4": "Senate 25",
    "dc49c0bc-8c68-11ed-b7db-fa163e33fba4": "District 24",
    "ddc66046-8c66-11ed-9adb-fa163ec0786c": "District 104",
    "e00dbeee-ec20-11ec-8b2b-fa163e761ca9": "treasurers list",
    "e033551a-8c60-11ed-a03e-fa163ec0786c": "Senate 63",
    "e5646e92-8c61-11ed-85e9-fa163ec0786c": "Senate 31",
    "e84c1156-8c65-11ed-a03e-fa163ec0786c": "District 75",
    "e989bfce-8c64-11ed-bb5c-fa163e7e5464": "District 39",
    "eef1c34a-8c68-11ed-bb5c-fa163e7e5464": "District 113",
    "ef2a3a32-8c62-11ed-a7e2-fa163e7e5464": "District 92",
    "f0f7d6f4-8c61-11ed-89c1-fa163e33fba4": "Senate 19",
    "f180aaac-2119-11f0-b4c1-fa163e7ee3ac": "Adding Project",
    "f29726dc-2301-11ed-83e1-fa163e5fbd10": "NYC Council",
    "f62a1664-8c66-11ed-85bd-fa163e761ca9": "District 25",
    "f8588cb2-eb45-11ec-a473-fa163e33fba4": "Assembly",
    "f87cd4f8-8c5f-11ed-988a-fa163e761ca9": "Senate 56",
    "fa8bd4a6-8c64-11ed-bb5c-fa163e7e5464": "District 43",
    "fac5ff02-eb45-11ec-a0ac-fa163e7e5464": "Senate",
    "fc22d970-8c61-11ed-9524-fa163ec0786c": "Senate 55",
}

# VIP-worthy tag IDs (for CC member filtering)
CC_VIP_TAG_IDS = set(CC_TAG_NAME_MAP.keys())


# ----------------------------
# Secrets helper
# ----------------------------
def _get_secret(key, default=""):
    try:
        val = st.secrets.get(key, default)
        return val if val else default
    except Exception:
        return default


# ----------------------------
# CC Token refresh
# ----------------------------
def _cc_refresh_token(refresh_token):
    """Exchange a CC refresh token for a fresh access token.
    Returns (access_token, expires_in) or raises on failure.
    """
    import base64 as _b64
    basic = _b64.b64encode(f"{CC_CLIENT_ID}:{CC_CLIENT_SECRET}".encode()).decode()
    resp = requests.post(
        CC_TOKEN_URL,
        data={
            "grant_type":    "refresh_token",
            "refresh_token": refresh_token,
        },
        headers={
            "Content-Type":  "application/x-www-form-urlencoded",
            "Authorization": f"Basic {basic}",
        },
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()
    return data["access_token"], data.get("expires_in", 28800)


def _cc_ensure_token():
    """Ensure a valid CC access token is in session_state.
    Called once per session (checks if already set to avoid re-refreshing).
    Sets st.session_state['cc_access_token'] or st.session_state['cc_error'].
    """
    if st.session_state.get("cc_access_token"):
        return  # already set this session

    refresh_token = _get_secret("CC_REFRESH_TOKEN")
    if refresh_token:
        try:
            access_token, _ = _cc_refresh_token(refresh_token)
            st.session_state["cc_access_token"] = access_token
            st.session_state.pop("cc_error", None)
            return
        except Exception as e:
            st.session_state["cc_error"] = f"Refresh failed: {e}"

    # Fallback: direct access token (for local dev)
    access_token = _get_secret("CC_ACCESS_TOKEN")
    if access_token:
        st.session_state["cc_access_token"] = access_token
        st.session_state.pop("cc_error", None)
    else:
        st.session_state["cc_error"] = (
            "No CC credentials. Set CC_REFRESH_TOKEN in Streamlit secrets."
        )


# ----------------------------
# MC Cache
# ----------------------------
def _mc_cache_save(members, campaigns, opener_sets):
    try:
        cache = {
            "saved_at": datetime.now().isoformat(),
            "members": members,
            "campaigns": campaigns,
            "opener_sets": [list(s) for s in opener_sets],
        }
        with open(MC_CACHE_PATH, "w") as f:
            json.dump(cache, f)
    except Exception:
        pass


def _mc_cache_load():
    """Load MC cache — tries local user cache first, then repo-bundled fallback."""
    for path in [MC_CACHE_PATH, MC_CACHE_REPO_PATH]:
        try:
            if not os.path.exists(path):
                continue
            with open(path) as f:
                cache = json.load(f)
            opener_sets = [set(s) for s in cache.get("opener_sets", [])]
            return (
                cache.get("members", []),
                cache.get("campaigns", []),
                opener_sets,
                cache.get("saved_at", "unknown"),
            )
        except Exception:
            continue
    return None


# ----------------------------
# HTTP helper
# ----------------------------
def _get_with_retry(url, headers, retries=3, backoff=1, timeout=15):
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
        except requests.exceptions.RequestException:
            if attempt < retries - 1:
                time.sleep(backoff * (2 ** attempt))
                continue
            raise
        if r.status_code in (200, 404):
            return r
        if r.status_code in (503, 429, 500):
            if attempt < retries - 1:
                if r.status_code == 429:
                    retry_after = int(r.headers.get("Retry-After", backoff * (2 ** attempt)))
                    time.sleep(min(retry_after, 60))
                else:
                    time.sleep(backoff * (2 ** attempt))
                continue
        return r
    return r


# ----------------------------
# Mailchimp API helpers
# ----------------------------
def mc_get_all_tagged_members(api_key, list_id):
    server = api_key.split("-")[-1]
    base = f"https://{server}.api.mailchimp.com/3.0"
    headers = {"Authorization": f"apikey {api_key}"}

    segments = []
    offset = 0
    while True:
        url = (f"{base}/lists/{list_id}/segments?type=static"
               f"&count=100&offset={offset}")
        r = _get_with_retry(url, headers, retries=3, timeout=30)
        r.raise_for_status()
        batch = r.json().get("segments", [])
        if not batch:
            break
        segments.extend(batch)
        if len(batch) < 100:
            break
        offset += 100

    targeted_segs = [
        s for s in segments
        if s["id"] not in MC_BULK_SEGMENT_IDS and s.get("member_count", 0) > 0
    ]

    member_emails = set()
    raw_members = {}
    consecutive_failures = 0
    for seg in targeted_segs:
        seg_id = seg["id"]
        seg_offset = 0
        while True:
            url = (f"{base}/lists/{list_id}/segments/{seg_id}/members"
                   f"?count=500&offset={seg_offset}")
            r = _get_with_retry(url, headers, retries=3, timeout=30)
            if r.status_code != 200:
                consecutive_failures += 1
                if consecutive_failures >= 3:
                    raise RuntimeError(f"MC API returning {r.status_code} consistently")
                break
            batch = r.json().get("members", [])
            for m in batch:
                email = m["email_address"]
                email_lower = email.lower()
                if email_lower not in member_emails:
                    member_emails.add(email_lower)
                    raw_members[email_lower] = m
            consecutive_failures = 0
            if len(batch) < 500:
                break
            seg_offset += 500

    if not raw_members:
        return []

    def fetch_tags(email_lower):
        m = raw_members[email_lower]
        email = m["email_address"]
        h = hashlib.md5(email.lower().encode()).hexdigest()
        try:
            r = _get_with_retry(
                f"{base}/lists/{list_id}/members/{h}?fields=email_address,merge_fields,tags",
                headers
            )
            if r.status_code == 200:
                data = r.json()
                tag_names = [
                    t["name"] for t in data.get("tags", [])
                    if t.get("name") != "9/8/2024"
                ]
                return {
                    "Email Address": email,
                    "First Name": data.get("merge_fields", {}).get("FNAME", "")
                                  or m.get("merge_fields", {}).get("FNAME", ""),
                    "Last Name": data.get("merge_fields", {}).get("LNAME", "")
                                 or m.get("merge_fields", {}).get("LNAME", ""),
                    "TAGS": ",".join(tag_names),
                }
        except Exception:
            pass
        return {
            "Email Address": email,
            "First Name": m.get("merge_fields", {}).get("FNAME", ""),
            "Last Name": m.get("merge_fields", {}).get("LNAME", ""),
            "TAGS": "VIP",
        }

    with concurrent.futures.ThreadPoolExecutor(max_workers=40) as executor:
        results = list(executor.map(fetch_tags, list(raw_members.keys())))

    return results


def mc_get_campaigns(api_key, list_id, start_date, end_date):
    server = api_key.split("-")[-1]
    base = f"https://{server}.api.mailchimp.com/3.0"
    headers = {"Authorization": f"apikey {api_key}"}

    campaigns = []
    offset = 0
    count = 200
    start_str = start_date.strftime("%Y-%m-%dT%H:%M:%S+00:00")
    end_str = end_date.strftime("%Y-%m-%dT%H:%M:%S+00:00")

    while True:
        url = (f"{base}/campaigns?status=sent&list_id={list_id}"
               f"&since_send_time={start_str}&before_send_time={end_str}"
               f"&count={count}&offset={offset}")
        r = _get_with_retry(url, headers, retries=3, timeout=30)
        r.raise_for_status()
        batch = r.json().get("campaigns", [])
        if not batch:
            break
        for c in batch:
            send_time = c.get("send_time", "")
            campaigns.append({
                "id": c["id"],
                "subject": c.get("settings", {}).get("subject_line", ""),
                "send_time": send_time,
                "date": send_time[:10] if send_time else "",
            })
        if len(batch) < count:
            break
        offset += count

    campaigns.sort(key=lambda x: x["date"])
    return campaigns


def mc_get_openers(api_key, campaign_id):
    server = api_key.split("-")[-1]
    base = f"https://{server}.api.mailchimp.com/3.0"
    headers = {"Authorization": f"apikey {api_key}"}
    count = 1000

    def fetch_page(offset):
        url = f"{base}/reports/{campaign_id}/open-details?count={count}&offset={offset}"
        r = _get_with_retry(url, headers, retries=3, timeout=20)
        if r.status_code == 404:
            return [], 0
        if r.status_code != 200:
            raise RuntimeError(f"MC openers API returned {r.status_code} for {campaign_id}")
        data = r.json()
        return data.get("members", []), data.get("total_opens", 0)

    first_batch, total_opens = fetch_page(0)
    if not first_batch:
        return set()

    openers = {m["email_address"].lower() for m in first_batch}

    remaining_pages = math.ceil(total_opens / count) - 1
    if remaining_pages > 0:
        remaining_offsets = [count * (i + 1) for i in range(remaining_pages)]
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as pool:
            futs = [pool.submit(fetch_page, off) for off in remaining_offsets]
            for fut in concurrent.futures.as_completed(futs):
                batch, _ = fut.result()
                for m in batch:
                    openers.add(m["email_address"].lower())

    return openers


# ----------------------------
# Constant Contact API helpers
# ----------------------------
def cc_get_all_tagged_members(access_token):
    base = "https://api.cc.email/v3"
    headers = {"Authorization": f"Bearer {access_token}"}

    tagged_members = []
    cursor = None
    while True:
        url = f"{base}/contacts?limit=500&include=taggings"
        if cursor:
            url += f"&cursor={cursor}"
        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()

        for c in data.get("contacts", []):
            taggings = c.get("taggings", [])
            if not any(tid in CC_TAG_NAME_MAP for tid in taggings):
                continue

            email_obj = c.get("email_address", {})
            email = email_obj.get("address", "") if isinstance(email_obj, dict) else ""
            if not email:
                continue

            tag_names = [CC_TAG_NAME_MAP[tid] for tid in taggings if tid in CC_TAG_NAME_MAP]
            tags_str = ",".join(tag_names) if tag_names else "VIP"

            tagged_members.append({
                "Email Address": email,
                "First Name": c.get("first_name", "") or "",
                "Last Name": c.get("last_name", "") or "",
                "TAGS": tags_str,
            })

        links = data.get("_links", {})
        next_href = links.get("next", {}).get("href", "") if isinstance(links.get("next"), dict) else ""
        if next_href and "cursor=" in next_href:
            cursor = next_href.split("cursor=")[-1].split("&")[0]
        else:
            break

    return tagged_members


def cc_get_campaigns(access_token, start_date, end_date):
    base = "https://api.cc.email/v3"
    headers = {"Authorization": f"Bearer {access_token}"}

    campaigns = []
    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    next_url = f"{base}/emails?status=SENT&limit=50"

    while next_url:
        r = requests.get(next_url, headers=headers, timeout=30)
        if r.status_code != 200:
            break

        data = r.json()
        page_campaigns = data.get("campaigns", [])
        if not page_campaigns:
            break

        stop_early = False
        for c in page_campaigns:
            if c.get("current_status", "").upper() != "DONE":
                continue
            created_at = c.get("created_at", "")
            if not created_at:
                continue
            date_str = created_at[:10]
            if date_str < start_str:
                stop_early = True
                break
            if date_str > end_str:
                continue

            campaign_id = c["campaign_id"]
            r2 = requests.get(f"{base}/emails/{campaign_id}", headers=headers, timeout=30)
            if r2.status_code != 200:
                continue
            act_data = r2.json()
            primary_act_id = None
            for act in act_data.get("campaign_activities", []):
                if act.get("role") == "primary_email":
                    primary_act_id = act["campaign_activity_id"]
                    break

            if primary_act_id:
                campaigns.append({
                    "id": primary_act_id,
                    "campaign_id": campaign_id,
                    "name": c.get("name", ""),
                    "created_at": created_at,
                    "date": date_str,
                })

        if stop_early:
            break

        links = data.get("_links", {})
        next_href = links.get("next", {}).get("href", "")
        if next_href:
            next_url = f"https://api.cc.email{next_href}"
        else:
            next_url = None

    campaigns.sort(key=lambda x: x["date"])
    return campaigns


def cc_get_openers(access_token, campaign_activity_id):
    base = "https://api.cc.email/v3"
    headers = {"Authorization": f"Bearer {access_token}"}

    openers = set()
    url = f"{base}/reports/email_reports/{campaign_activity_id}/tracking/unique_opens?limit=500"

    while True:
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            break
        data = r.json()
        for contact in data.get("tracking_activities", []):
            email = contact.get("email_address", "")
            if email:
                openers.add(email.lower())
        links = data.get("_links", {})
        next_href = links.get("next", {}).get("href", "")
        if next_href:
            url = f"https://api.cc.email{next_href}"
        else:
            break

    return openers


# ----------------------------
# Build report DataFrames
# ----------------------------
def build_reports(mc_members, cc_members,
                  mc_campaigns, cc_campaigns,
                  mc_opener_sets, cc_opener_sets):
    cc_date_openers = {}
    for camp, openers in zip(cc_campaigns, cc_opener_sets):
        date = camp["date"]
        if date:
            cc_date_openers[date] = openers

    mc_date_openers = {}
    for camp, openers in zip(mc_campaigns, mc_opener_sets):
        date = camp["date"]
        if date:
            mc_date_openers[date] = openers

    all_dates = sorted(set(list(cc_date_openers.keys()) + list(mc_date_openers.keys())))

    def build_esp_df(members, date_openers, esp_name):
        if not members:
            return pd.DataFrame()

        df = pd.DataFrame(members)
        if "TAGS" not in df.columns:
            df["TAGS"] = ""
        df.set_index("Email Address", inplace=True)

        for i, date in enumerate(all_dates, start=1):
            col_name = f"Campaign {i}"
            openers = date_openers.get(date, set())
            df[col_name] = df.index.map(lambda e: "X" if e.lower() in openers else None)

        campaign_cols = [f"Campaign {i}" for i in range(1, len(all_dates) + 1)]
        campaign_cols = [c for c in campaign_cols if c in df.columns]

        if campaign_cols:
            df["Total Opens"] = df[campaign_cols].apply(
                lambda row: (row == "X").sum(), axis=1
            )
        else:
            df["Total Opens"] = 0

        df["ESP"] = esp_name
        df.reset_index(inplace=True)

        fixed_order = ["Email Address", "First Name", "Last Name", "TAGS", "Total Opens"]
        other_cols = [c for c in df.columns if c not in fixed_order + ["ESP"]]
        df = df[fixed_order + other_cols + ["ESP"]]
        return df

    cc_df = build_esp_df(cc_members, cc_date_openers, "Constant Contact")
    mc_df = build_esp_df(mc_members, mc_date_openers, "Mailchimp")

    return cc_df, mc_df, all_dates


# ----------------------------
# Excel export
# ----------------------------
def build_excel(mc_df, cc_df, all_dates=None):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    CC_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")
    MC_ROW_FILL  = PatternFill("solid", fgColor="E8F5E9")
    CC_ESP_FILL  = PatternFill("solid", fgColor="2980B9")
    MC_ESP_FILL  = PatternFill("solid", fgColor="27AE60")
    X_FILL       = PatternFill("solid", fgColor="C8E6C9")
    HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
    LOOKUP_HDR   = PatternFill("solid", fgColor="1A5276")
    SUMM_HDR     = PatternFill("solid", fgColor="1B2631")
    SUMM_SECTION = PatternFill("solid", fgColor="2E4057")
    SUMM_CC      = PatternFill("solid", fgColor="D6E4F0")
    SUMM_MC      = PatternFill("solid", fgColor="E8F5E9")
    SUMM_ALT     = PatternFill("solid", fgColor="F2F3F4")

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ESP_FONT     = Font(bold=True, color="FFFFFF", size=10)
    X_FONT       = Font(bold=True, color="1B5E20", size=10)
    SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(bold=True, color="FFFFFF", size=14)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    WRAP_LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="BDBDBD")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    FIXED = {"Email Address", "First Name", "Last Name", "TAGS", "Total Opens", "ESP"}

    camp_label_map = {}
    if all_dates:
        for i, d in enumerate(all_dates, start=1):
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                camp_label_map[f"Campaign {i}"] = dt.strftime("%m/%d")
            except Exception:
                camp_label_map[f"Campaign {i}"] = f"Campaign {i}"

    output = BytesIO()

    dfs_to_combine = []
    if cc_df is not None and not cc_df.empty:
        dfs_to_combine.append(cc_df)
    if mc_df is not None and not mc_df.empty:
        dfs_to_combine.append(mc_df)

    if not dfs_to_combine:
        output.seek(0)
        return output

    master_df = pd.concat(dfs_to_combine, ignore_index=True)
    master_df = master_df.sort_values("Total Opens", ascending=False).reset_index(drop=True)

    def _format_sheet(ws, df, rename_camps=True):
        if rename_camps and camp_label_map:
            for cell in ws[1]:
                if cell.value in camp_label_map:
                    cell.value = camp_label_map[cell.value]

        headers = [c.value for c in ws[1]]
        n_cols  = len(headers)
        n_rows  = ws.max_row

        def _ci(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        esp_ci   = _ci("ESP")
        opens_ci = _ci("Total Opens")
        tags_ci  = _ci("TAGS")
        email_ci = _ci("Email Address")

        camp_col_set = set(range(1, n_cols + 1)) - {
            c for c in range(1, n_cols + 1)
            if headers[c-1] in FIXED
        }

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = CENTER if col_idx in camp_col_set or col_idx == opens_ci else LEFT
            cell.border    = BORDER

        for row in range(2, n_rows + 1):
            ws.row_dimensions[row].height = 18
            esp_val  = ws.cell(row, esp_ci).value if esp_ci else None
            row_fill = CC_ROW_FILL if esp_val == "Constant Contact" else MC_ROW_FILL

            for col in range(1, n_cols + 1):
                cell = ws.cell(row, col)
                cell.border = BORDER
                if col == esp_ci:
                    cell.fill      = CC_ESP_FILL if esp_val == "Constant Contact" else MC_ESP_FILL
                    cell.font      = ESP_FONT
                    cell.alignment = CENTER
                elif col in camp_col_set and cell.value == "X":
                    cell.fill      = X_FILL
                    cell.font      = X_FONT
                    cell.alignment = CENTER
                elif col in camp_col_set:
                    cell.fill      = row_fill
                    cell.alignment = CENTER
                elif col == opens_ci:
                    cell.fill      = row_fill
                    cell.alignment = CENTER
                elif col == tags_ci:
                    cell.fill      = row_fill
                    cell.alignment = WRAP_LEFT
                else:
                    cell.fill      = row_fill
                    cell.alignment = LEFT

        for col in range(1, n_cols + 1):
            header = headers[col - 1] or ""
            if col == email_ci:
                w = 30
            elif header in ("First Name", "Last Name"):
                w = 15
            elif col == tags_ci:
                w = 22
            elif col == esp_ci:
                w = 18
            elif col == opens_ci:
                w = 13
            elif col in camp_col_set:
                w = 9
            else:
                w = 14
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

    def _write_summary_sheet(ws, master):
        ws.sheet_properties.tabColor = "1B2631"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18

        row = 1

        def _title(text, fill=SUMM_HDR):
            ws.row_dimensions[row].height = 30
            c = ws.cell(row, 1, text)
            c.fill = fill; c.font = TITLE_FONT; c.alignment = LEFT
            ws.merge_cells(f"A{row}:C{row}")
            return row

        def _section(text):
            nonlocal row
            ws.row_dimensions[row].height = 22
            c = ws.cell(row, 1, text)
            c.fill = SUMM_SECTION; c.font = SECTION_FONT; c.alignment = LEFT
            ws.merge_cells(f"A{row}:C{row}")
            row += 1

        def _header_row(cols):
            nonlocal row
            ws.row_dimensions[row].height = 20
            for ci, label in enumerate(cols, start=1):
                c = ws.cell(row, ci, label)
                c.fill = SUMM_HDR; c.font = HEADER_FONT
                c.alignment = CENTER if ci > 1 else LEFT
                c.border = BORDER
            row += 1

        def _data_row(vals, fill=None, bold=False):
            nonlocal row
            ws.row_dimensions[row].height = 18
            alt = SUMM_ALT if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row, ci, v)
                c.fill     = fill or alt
                c.font     = Font(bold=bold, size=10)
                c.alignment = CENTER if ci > 1 else LEFT
                c.border   = BORDER
            row += 1

        _title("  VIP Report — Summary")
        row += 1

        _section("  Subscriber Totals")
        _header_row(["ESP", "Subscribers", ""])
        cc_count = int((master["ESP"] == "Constant Contact").sum())
        mc_count = int((master["ESP"] == "Mailchimp").sum())
        _data_row(["Constant Contact", cc_count, ""], fill=SUMM_CC)
        _data_row(["Mailchimp",        mc_count, ""], fill=SUMM_MC)
        _data_row(["Total",            cc_count + mc_count, ""],
                  fill=PatternFill("solid", fgColor="D5D8DC"), bold=True)
        row += 1

        _section("  Opens per Campaign")
        _header_row(["Campaign", "Date", "Total Openers"])

        camp_orig_cols = [c for c in master.columns if c not in FIXED]
        for i, col in enumerate(camp_orig_cols, start=1):
            label    = camp_label_map.get(col, col)
            n_openers = int((master[col] == "X").sum())
            alt = SUMM_ALT if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            _data_row([f"Campaign {i}", label, n_openers], fill=alt)
        row += 1

        _section("  Top 10 Openers (all campaigns)")
        _header_row(["Name", "Email", "Total Opens"])

        top10 = master.nlargest(10, "Total Opens")
        for _, r in top10.iterrows():
            name     = f"{r.get('First Name', '')} {r.get('Last Name', '')}".strip()
            esp_val  = r.get("ESP", "")
            fill     = CC_ROW_FILL if esp_val == "Constant Contact" else MC_ROW_FILL
            _data_row([name, r.get("Email Address", ""), int(r.get("Total Opens", 0))],
                      fill=fill)

        ws.freeze_panes = "A2"

    def _write_lookup_sheet(ws, df):
        ws.sheet_properties.tabColor = "1A5276"
        lookup_cols = ["Email Address", "First Name", "Last Name", "TAGS", "Total Opens", "ESP"]
        lookup_cols = [c for c in lookup_cols if c in df.columns]
        ldf = df[lookup_cols].copy()

        ws.row_dimensions[1].height = 28
        for ci, col_name in enumerate(lookup_cols, start=1):
            cell = ws.cell(1, ci, col_name)
            cell.fill = LOOKUP_HDR; cell.font = HEADER_FONT
            cell.alignment = LEFT; cell.border = BORDER

        for ri, (_, row_data) in enumerate(ldf.iterrows(), start=2):
            ws.row_dimensions[ri].height = 18
            esp_val  = row_data.get("ESP", "")
            row_fill = CC_ROW_FILL if esp_val == "Constant Contact" else MC_ROW_FILL

            for ci, col_name in enumerate(lookup_cols, start=1):
                cell = ws.cell(ri, ci, row_data[col_name])
                cell.border = BORDER
                if col_name == "ESP":
                    cell.fill = CC_ESP_FILL if esp_val == "Constant Contact" else MC_ESP_FILL
                    cell.font = ESP_FONT; cell.alignment = CENTER
                elif col_name == "Total Opens":
                    cell.fill = row_fill; cell.alignment = CENTER
                else:
                    cell.fill = row_fill; cell.alignment = LEFT

        widths = {"Email Address": 32, "First Name": 15, "Last Name": 15,
                  "TAGS": 26, "Total Opens": 13, "ESP": 18}
        for ci, col_name in enumerate(lookup_cols, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = widths.get(col_name, 14)

        n_rows = ldf.shape[0] + 1
        ws.auto_filter.ref = f"A1:{get_column_letter(len(lookup_cols))}{n_rows}"
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        master_df.to_excel(writer, sheet_name="All", index=False)
        _format_sheet(writer.sheets["All"], master_df)

        if cc_df is not None and not cc_df.empty:
            cc_sorted = cc_df.sort_values("Total Opens", ascending=False).reset_index(drop=True)
            cc_sorted.to_excel(writer, sheet_name="Constant Contact", index=False)
            _format_sheet(writer.sheets["Constant Contact"], cc_sorted)

        if mc_df is not None and not mc_df.empty:
            mc_sorted = mc_df.sort_values("Total Opens", ascending=False).reset_index(drop=True)
            mc_sorted.to_excel(writer, sheet_name="Mailchimp", index=False)
            _format_sheet(writer.sheets["Mailchimp"], mc_sorted)

        lookup_ws = writer.book.create_sheet("Lookup")
        _write_lookup_sheet(lookup_ws, master_df)

        summ_ws = writer.book.create_sheet("Summary", 0)
        _write_summary_sheet(summ_ws, master_df)
        writer.book.active = summ_ws

    output.seek(0)
    return output


# ----------------------------
# Streamlit App
# ----------------------------
def main():
    st.set_page_config(
        page_title="VIP Report",
        layout="wide",
        page_icon="📊",
    )

    # Auto-refresh CC token once per session (silent, fast)
    _cc_ensure_token()

    st.title("VIP Report Generator")
    st.caption(
        "Pulls all tagged subscribers and campaign opens from Mailchimp + Constant Contact. "
        "Covers VIP, Hochul, Media, Assembly, Senate, and all other tagged contacts."
    )

    # --- Sidebar ---
    with st.sidebar:
        st.header("Report Settings")

        default_end   = datetime.today() - timedelta(days=1)
        default_start = default_end - timedelta(days=13)
        start_date = st.date_input("Start Date", value=default_start)
        end_date   = st.date_input("End Date",   value=default_end)

        st.divider()

        # Connection status
        mc_key  = _get_secret("MAILCHIMP_API_KEY")
        cc_tok  = st.session_state.get("cc_access_token", "")
        cc_err  = st.session_state.get("cc_error", "")

        st.markdown("**API Status**")
        st.markdown(f"{'🟢' if mc_key  else '🔴'} Mailchimp")
        st.markdown(f"{'🟢' if cc_tok  else '🔴'} Constant Contact" +
                    (f" — {cc_err}" if cc_err and not cc_tok else ""))

        if cc_err and not cc_tok:
            if st.button("Retry CC Connection"):
                st.session_state.pop("cc_access_token", None)
                st.session_state.pop("cc_error", None)
                st.rerun()

        st.divider()
        run_mc = st.checkbox("Include Mailchimp",         value=bool(mc_key))
        run_cc = st.checkbox("Include Constant Contact",  value=bool(cc_tok))

        st.divider()
        st.caption("MC: all static segments (~462 contacts)")
        st.caption("CC: tagged contacts (~791 contacts)")
        st.caption("Runtime: ~4–5 min for both ESPs")

    # --- Main: generate ---
    generate = st.button("Generate Report", type="primary", use_container_width=True)

    if generate:
        if not run_mc and not run_cc:
            st.error("Select at least one ESP in the sidebar.")
            return
        if run_cc and not cc_tok:
            st.error("Constant Contact token unavailable. Check sidebar for CC error.")
            return

        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt   = datetime.combine(end_date,   datetime.max.time().replace(microsecond=0))

        mc_members, mc_campaigns, mc_opener_sets = [], [], []
        cc_members, cc_campaigns, cc_opener_sets = [], [], []

        # --- Mailchimp ---
        if run_mc:
            mc_api_key = _get_secret("MAILCHIMP_API_KEY")
            mc_list_id = _get_secret("MAILCHIMP_LIST_ID")
            if not mc_api_key or not mc_list_id:
                st.error("Mailchimp credentials missing from secrets.")
            else:
                mc_cache_used = False
                with st.spinner("Fetching Mailchimp tagged members (~1–2 min)..."):
                    try:
                        mc_members = mc_get_all_tagged_members(mc_api_key, mc_list_id)
                        st.success(f"Mailchimp: {len(mc_members)} tagged members found.")
                    except Exception as e:
                        cached = _mc_cache_load()
                        if cached:
                            mc_members, _cc, _co, saved_at = cached
                            start_str = start_dt.strftime("%Y-%m-%d")
                            end_str   = end_dt.strftime("%Y-%m-%d")
                            filtered       = [(c, o) for c, o in zip(_cc, _co)
                                              if start_str <= c.get("date", "") <= end_str]
                            mc_campaigns   = [c for c, _ in filtered]
                            mc_opener_sets = [o for _, o in filtered]
                            st.warning(
                                f"Mailchimp API unavailable ({e.__class__.__name__}). "
                                f"Using cached data from {saved_at[:16]}."
                            )
                            mc_cache_used = True
                        else:
                            st.error(f"Mailchimp fetch failed and no cache available: {e}")

                if mc_members and not mc_cache_used:
                    with st.spinner("Fetching Mailchimp campaigns..."):
                        try:
                            mc_campaigns = mc_get_campaigns(mc_api_key, mc_list_id, start_dt, end_dt)
                            st.info(f"Mailchimp: {len(mc_campaigns)} campaigns in range.")
                        except Exception as e:
                            st.error(f"MC campaign fetch failed: {e}")

                    for camp in mc_campaigns:
                        label = camp.get("date", "") or camp["id"]
                        with st.spinner(f"MC openers: {label}..."):
                            try:
                                mc_opener_sets.append(mc_get_openers(mc_api_key, camp["id"]))
                            except Exception as e:
                                st.warning(f"MC openers {label}: {e}")
                                mc_opener_sets.append(set())

                    _mc_cache_save(mc_members, mc_campaigns, mc_opener_sets)

        # --- Constant Contact ---
        if run_cc:
            access_token = st.session_state.get("cc_access_token", "")
            with st.spinner("Scanning CC contacts for tagged members (~2–3 min)..."):
                try:
                    cc_members = cc_get_all_tagged_members(access_token)
                    st.success(f"Constant Contact: {len(cc_members)} tagged members found.")
                except Exception as e:
                    if "401" in str(e) or "Unauthorized" in str(e):
                        # Token may have expired mid-session — try one refresh
                        refresh_token = _get_secret("CC_REFRESH_TOKEN")
                        if refresh_token:
                            try:
                                new_token, _ = _cc_refresh_token(refresh_token)
                                st.session_state["cc_access_token"] = new_token
                                cc_members = cc_get_all_tagged_members(new_token)
                                st.success(
                                    f"Constant Contact (re-authed): {len(cc_members)} tagged members found."
                                )
                                access_token = new_token
                            except Exception as e2:
                                st.error(f"CC token refresh failed: {e2}")
                        else:
                            st.error("CC token expired. Contact Nick to refresh credentials.")
                    else:
                        st.error(f"CC fetch failed: {e}")

            if cc_members:
                with st.spinner("Fetching CC campaigns..."):
                    try:
                        cc_campaigns = cc_get_campaigns(access_token, start_dt, end_dt)
                        st.info(f"Constant Contact: {len(cc_campaigns)} campaigns in range.")
                    except Exception as e:
                        st.error(f"CC campaign fetch failed: {e}")

                for camp in cc_campaigns:
                    label = camp.get("date", "") or camp["id"]
                    with st.spinner(f"CC openers: {label}..."):
                        try:
                            cc_opener_sets.append(cc_get_openers(access_token, camp["id"]))
                        except Exception as e:
                            st.warning(f"CC openers {label}: {e}")
                            cc_opener_sets.append(set())

        # Build report
        cc_df, mc_df, all_dates = build_reports(
            mc_members, cc_members,
            mc_campaigns, cc_campaigns,
            mc_opener_sets, cc_opener_sets,
        )

        st.session_state.update({
            "mc_df": mc_df,
            "cc_df": cc_df,
            "all_dates": all_dates,
            "report_generated": True,
        })

    # --- Display results ---
    if st.session_state.get("report_generated"):
        mc_df     = st.session_state.get("mc_df")
        cc_df     = st.session_state.get("cc_df")
        all_dates = st.session_state.get("all_dates", [])

        tab1, tab2 = st.tabs(["Dashboard", "Download Excel"])

        with tab1:
            fixed_cols = {"Email Address", "First Name", "Last Name", "TAGS", "Total Opens", "ESP"}

            total_vips = sum([
                len(mc_df) if mc_df is not None and not mc_df.empty else 0,
                len(cc_df) if cc_df is not None and not cc_df.empty else 0,
            ])

            col1, col2, col3 = st.columns(3)
            col1.metric("Total Tagged Subscribers", total_vips)
            col2.metric("Campaigns (date range)", len(all_dates))
            col3.metric(
                "Date Range",
                f"{all_dates[0][:5] if all_dates else '—'} → {all_dates[-1][:5] if all_dates else '—'}"
            )

            if all_dates:
                st.caption(f"{all_dates[0]} → {all_dates[-1]}  ({len(all_dates)} issues)")

            st.divider()

            if cc_df is not None and not cc_df.empty:
                st.subheader(f"Constant Contact — {len(cc_df):,} contacts")
                camp_subset = [c for c in cc_df.columns if c not in fixed_cols]
                st.dataframe(
                    cc_df.style.map(
                        lambda v: "background-color: #d4edda; color: #155724;" if v == "X" else "",
                        subset=camp_subset if camp_subset else []
                    ),
                    use_container_width=True,
                )

            if mc_df is not None and not mc_df.empty:
                st.subheader(f"Mailchimp — {len(mc_df):,} contacts")
                camp_subset = [c for c in mc_df.columns if c not in fixed_cols]
                st.dataframe(
                    mc_df.style.map(
                        lambda v: "background-color: #d4edda; color: #155724;" if v == "X" else "",
                        subset=camp_subset if camp_subset else []
                    ),
                    use_container_width=True,
                )

        with tab2:
            st.subheader("Download Excel Report")
            if (cc_df is not None and not cc_df.empty) or (mc_df is not None and not mc_df.empty):
                excel_data = build_excel(mc_df, cc_df, all_dates=all_dates)
                today = datetime.today().strftime("%m_%d_%y")
                st.download_button(
                    label="Download VIP Report (.xlsx)",
                    data=excel_data,
                    file_name=f"VIP Report {today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
            else:
                st.info("No data to export. Generate the report first.")


if __name__ == "__main__":
    main()
